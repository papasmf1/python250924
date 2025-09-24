import requests
from bs4 import BeautifulSoup
from urllib.parse import quote
from openpyxl import Workbook
from datetime import datetime

def get_news_titles(search_keyword):
    # 검색어 URL 인코딩
    encoded_keyword = quote(search_keyword)
    
    # 검색 URL 생성
    url = f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={encoded_keyword}"
    
    # User-Agent 설정
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    
    # 웹 페이지 요청
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        # HTML 파싱
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 뉴스 제목 요소 찾기 (클래스명 참조)
        news_titles = soup.select(".sds-comps-text-type-headline1")
        
        # 제목 텍스트 추출 및 저장
        titles = []
        for title in news_titles:
            # HTML 마크업 제거 후 텍스트만 추출
            clean_title = title.get_text().strip()
            titles.append(clean_title)
            
        return titles
    else:
        print(f"Error: {response.status_code}")
        return []

def save_to_excel(keyword, titles):
    # 새로운 워크북 생성
    wb = Workbook()
    ws = wb.active
    
    # 워크시트 제목 설정
    ws.title = "네이버 뉴스 검색결과"
    
    # 헤더 추가
    ws['A1'] = '검색어'
    ws['B1'] = '검색 날짜'
    ws['C1'] = '번호'
    ws['D1'] = '뉴스 제목'
    
    # 검색어와 날짜 정보 추가
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A2'] = keyword
    ws['B2'] = current_date
    
    # 뉴스 제목 추가
    for idx, title in enumerate(titles, 1):
        ws[f'C{idx+1}'] = idx
        ws[f'D{idx+1}'] = title
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 파일 저장
    filename = f'results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    print(f'\nExcel 파일이 생성되었습니다: {filename}')

if __name__ == "__main__":
    # 검색어 설정
    search_keyword = "아이폰17"
    
    # 뉴스 제목 가져오기
    titles = get_news_titles(search_keyword)
    
    # 결과 출력
    print(f"'{search_keyword}' 관련 뉴스 제목:")
    print("-" * 50)
    for i, title in enumerate(titles, 1):
        print(f"{i}. {title}")
    
    # Excel 파일로 저장
    save_to_excel(search_keyword, titles)