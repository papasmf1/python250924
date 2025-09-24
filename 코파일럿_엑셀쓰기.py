#pip install faker 
import random
from openpyxl import Workbook
from faker import Faker

# Faker를 한국어로 설정
fake = Faker('ko_KR')

# 전자제품 목록 정의
electronic_products = [
    "스마트폰", "노트북", "태블릿", "스마트워치", "무선이어폰", 
    "블루투스 스피커", "공기청정기", "로봇청소기", "전기밥솥", "전자레인지",
    "냉장고", "세탁기", "건조기", "TV", "모니터"
]

# 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "제품 데이터"

# 헤더 추가
headers = ["제품ID", "제품명", "수량", "가격"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 100개의 데이터 생성
for row in range(2, 102):
    product = random.choice(electronic_products)
    
    # 데이터 생성:문자열 포맷팅 
    product_id = f"P{str(row-1).zfill(3)}"  # P001, P002, ...
    quantity = random.randint(1, 100)
    
    # 제품별 가격 범위 설정
    if "스마트폰" in product or "노트북" in product:
        price = random.randint(800000, 2000000)
    elif "TV" in product or "냉장고" in product:
        price = random.randint(500000, 1500000)
    else:
        price = random.randint(50000, 500000)
    
    # 데이터 입력
    ws.cell(row=row, column=1, value=product_id)
    ws.cell(row=row, column=2, value=product)
    ws.cell(row=row, column=3, value=quantity)
    ws.cell(row=row, column=4, value=price)

# 파일 저장
wb.save('products.xlsx')
print("데이터가 성공적으로 저장되었습니다.")